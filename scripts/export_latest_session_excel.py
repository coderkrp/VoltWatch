from __future__ import annotations

import argparse
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB_PATH = REPO_ROOT / "backend" / "data_logger.db"
DEFAULT_OUTPUT_PATH = REPO_ROOT / "exports" / "all_sessions_export.xlsx"


def _sanitize(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._-") or "value"


def _sheet_name(session_id: int, sensor_index: int, used_names: set[str]) -> str:
    base_name = f"s{session_id}_sensor{sensor_index}"
    candidate = base_name[:31]
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    used_names.add(candidate)
    return candidate


def _session_sensor_rows(conn: sqlite3.Connection, device_id: str | None) -> list[sqlite3.Row]:
    params: list[str] = []
    query = """
        SELECT
            sess.session_id,
            sess.device_id,
            sess.start_time,
            sess.end_time,
            sensor.sensor_id,
            sensor.sensor_index,
            COALESCE(sensor.alias, '') AS sensor_alias,
            COUNT(reading.id) AS reading_count
        FROM sessions AS sess
        JOIN readings AS reading
            ON reading.session_id = sess.session_id
        JOIN sensors AS sensor
            ON sensor.sensor_id = reading.sensor_id
        WHERE 1 = 1
    """
    if device_id:
        query += " AND sess.device_id = ?"
        params.append(device_id)

    query += """
        GROUP BY
            sess.session_id,
            sess.device_id,
            sess.start_time,
            sess.end_time,
            sensor.sensor_id,
            sensor.sensor_index,
            sensor.alias
        ORDER BY sess.start_time, sensor.sensor_index
    """
    return conn.execute(query, params).fetchall()


def _reading_rows(conn: sqlite3.Connection, session_id: int, sensor_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT
            seq,
            timestamp_esp,
            timestamp_server,
            bus_voltage,
            shunt_voltage,
            supply_voltage,
            current,
            supply_voltage * current AS power
        FROM readings
        WHERE session_id = ? AND sensor_id = ?
        ORDER BY timestamp_esp, seq
        """,
        (session_id, sensor_id),
    ).fetchall()


def _write_sheet(
    workbook: Workbook,
    session_sensor_row: sqlite3.Row,
    readings: list[sqlite3.Row],
    used_names: set[str],
) -> None:
    sheet = workbook.create_sheet(
        title=_sheet_name(
            session_id=session_sensor_row["session_id"],
            sensor_index=session_sensor_row["sensor_index"],
            used_names=used_names,
        )
    )

    metadata = [
        ("device_id", session_sensor_row["device_id"]),
        ("session_id", session_sensor_row["session_id"]),
        ("session_start_time", session_sensor_row["start_time"]),
        ("session_end_time", session_sensor_row["end_time"] or ""),
        ("sensor_id", session_sensor_row["sensor_id"]),
        ("sensor_index", session_sensor_row["sensor_index"]),
        ("sensor_alias", session_sensor_row["sensor_alias"]),
        ("row_count", len(readings)),
    ]
    for key, value in metadata:
        sheet.append([key, value])

    sheet.append([])
    sheet.append(
        [
            "seq",
            "timestamp_esp",
            "timestamp_server",
            "bus_voltage",
            "shunt_voltage",
            "supply_voltage",
            "current",
            "power",
        ]
    )

    for row in readings:
        sheet.append(
            [
                row["seq"],
                row["timestamp_esp"],
                row["timestamp_server"],
                row["bus_voltage"],
                row["shunt_voltage"],
                row["supply_voltage"],
                row["current"],
                row["power"],
            ]
        )


def export_all_session_sensor_sheets(
    db_path: Path,
    output_path: Path,
    device_id: str | None,
) -> tuple[Path, int]:
    if not db_path.exists():
        raise SystemExit(f"Database file not found: {db_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        session_sensor_rows = _session_sensor_rows(conn, device_id)
        if not session_sensor_rows:
            if device_id:
                raise SystemExit(f"No session/sensor data found for device_id={device_id!r}.")
            raise SystemExit("No session/sensor data found in the database.")

        workbook = Workbook()
        workbook.remove(workbook.active)
        used_names: set[str] = set()

        for row in session_sensor_rows:
            readings = _reading_rows(conn, row["session_id"], row["sensor_id"])
            if readings:
                _write_sheet(workbook, row, readings, used_names)

        if not workbook.sheetnames:
            raise SystemExit("No reading rows found in the database.")

        workbook.save(output_path)
        return output_path, len(workbook.sheetnames)
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export one workbook with a sheet for every session/sensor combination."
    )
    parser.add_argument(
        "--db-path",
        type=Path,
        default=DEFAULT_DB_PATH,
        help=f"Path to the SQLite database (default: {DEFAULT_DB_PATH})",
    )
    parser.add_argument(
        "--output-path",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Workbook output path (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--device-id",
        type=str,
        default=None,
        help="Optional device_id to limit the export to one device.",
    )
    args = parser.parse_args()

    output_path, sheet_count = export_all_session_sensor_sheets(
        db_path=args.db_path.resolve(),
        output_path=args.output_path.resolve(),
        device_id=args.device_id,
    )

    print(f"Exported workbook with {sheet_count} sheet(s):")
    print(output_path)


if __name__ == "__main__":
    main()
