import sys
import unittest
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from export_excel import (  # noqa: E402
    READINGS_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    build_session_workbook,
    export_filename,
    readings_export_frame,
)


class ExportExcelTests(unittest.TestCase):
    def test_readings_export_frame_adds_power_and_sorted_timestamp_strings(self) -> None:
        raw_df = pd.DataFrame(
            [
                {
                    "timestamp": "2026-04-04T10:00:05Z",
                    "bus_voltage": 4.2,
                    "shunt_voltage": 1.1,
                    "supply_voltage": 4.4,
                    "current": 100.0,
                },
                {
                    "timestamp": "2026-04-04T10:00:00Z",
                    "bus_voltage": 4.0,
                    "shunt_voltage": 1.0,
                    "supply_voltage": 4.2,
                    "current": 90.0,
                },
            ]
        )

        export_df = readings_export_frame(raw_df)

        self.assertEqual(
            list(export_df.columns),
            ["timestamp_utc", "bus_voltage", "shunt_voltage", "supply_voltage", "current", "power"],
        )
        self.assertEqual(export_df.iloc[0]["timestamp_utc"], "2026-04-04T10:00:00+00:00")
        self.assertAlmostEqual(export_df.iloc[1]["power"], 440.0)

    def test_build_session_workbook_contains_summary_and_readings_sheets(self) -> None:
        raw_df = pd.DataFrame(
            [
                {
                    "timestamp": "2026-04-04T10:00:00Z",
                    "bus_voltage": 4.0,
                    "shunt_voltage": 1.0,
                    "supply_voltage": 4.2,
                    "current": 90.0,
                }
            ]
        )

        workbook_bytes = build_session_workbook(
            device_id="dev-1",
            session_row={
                "session_id": 10,
                "start_time": "2026-04-04T09:55:00Z",
                "end_time": "2026-04-04T10:05:00Z",
            },
            sensor_row={
                "sensor_id": 7,
                "sensor_index": 0,
                "alias": "Main Sensor",
            },
            raw_df=raw_df,
            summary_stats={"Power Avg": 378.0, "Power Max": 378.0},
        )

        self.assertGreater(len(workbook_bytes), 0)

        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        self.assertEqual(workbook.sheetnames, [SUMMARY_SHEET_NAME, READINGS_SHEET_NAME])

        summary_sheet = workbook[SUMMARY_SHEET_NAME]
        summary_values = {
            row[0]: row[1]
            for row in summary_sheet.iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        }
        self.assertEqual(summary_values["device_id"], "dev-1")
        self.assertEqual(summary_values["session_id"], 10)
        self.assertEqual(summary_values["sensor_id"], 7)
        self.assertEqual(summary_values["row_count"], 1)
        self.assertEqual(summary_values["Power Avg"], 378.0)

        readings_sheet = workbook[READINGS_SHEET_NAME]
        header = [cell for cell in next(readings_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        self.assertEqual(
            header,
            ["timestamp_utc", "bus_voltage", "shunt_voltage", "supply_voltage", "current", "power"],
        )
        first_data_row = next(readings_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertEqual(first_data_row[0], "2026-04-04T10:00:00+00:00")
        self.assertEqual(first_data_row[-1], 378.0)

    def test_build_session_workbook_handles_empty_readings(self) -> None:
        workbook_bytes = build_session_workbook(
            device_id="dev-1",
            session_row={"session_id": 10, "start_time": None, "end_time": None},
            sensor_row={"sensor_id": 7, "sensor_index": 0, "alias": None},
            raw_df=pd.DataFrame(),
            summary_stats=None,
        )

        workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        readings_sheet = workbook[READINGS_SHEET_NAME]
        rows = list(readings_sheet.iter_rows(values_only=True))
        self.assertEqual(
            rows[0],
            ("timestamp_utc", "bus_voltage", "shunt_voltage", "supply_voltage", "current", "power"),
        )
        self.assertEqual(len(rows), 1)

    def test_export_filename_is_deterministic(self) -> None:
        self.assertEqual(export_filename(12, 4), "session_12_sensor_4.xlsx")


if __name__ == "__main__":
    unittest.main()
